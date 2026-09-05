"""Ergebnisse sichern - neun Wege.

CSV, Excel, JSON, JSONL, Markdown, HTML, PDF, Text und SQLite. Schwere
Abhängigkeiten (openpyxl, reportlab) sind optional: fehlt eine, meldet das
Programm das sauber, statt beim Start zu scheitern.
"""

from __future__ import annotations

import csv
import json
import sqlite3
import time
from dataclasses import dataclass
from html import escape
from pathlib import Path
from typing import Any, Callable

from .extractors import all_columns

Rows = list[dict[str, Any]]


@dataclass
class Format:
    """Ein Ausgabeformat mit Endung und Beschreibung."""

    key: str
    label: str
    suffix: str
    description: str
    needs: str = ""  # Name eines optionalen Pakets


FORMATS: list[Format] = [
    Format("csv", "CSV-Tabelle", ".csv", "Für Excel, Numbers, Google Tabellen"),
    Format("xlsx", "Excel-Mappe", ".xlsx", "Formatiert, mit Kopfzeile", needs="openpyxl"),
    Format("json", "JSON", ".json", "Für Weiterverarbeitung im Code"),
    Format("jsonl", "JSON Lines", ".jsonl", "Ein Datensatz je Zeile, gut für grosse Mengen"),
    Format("markdown", "Markdown", ".md", "Für Notizen, GitHub, Obsidian"),
    Format("html", "HTML-Seite", ".html", "Fertige Tabelle zum Ansehen im Browser"),
    Format("pdf", "PDF-Bericht", ".pdf", "Zum Weitergeben und Archivieren", needs="reportlab"),
    Format("txt", "Textdatei", ".txt", "Reiner Text ohne Formatierung"),
    Format("sqlite", "SQLite-Datenbank", ".db", "Abfragbar mit SQL"),
]

FORMAT_BY_KEY = {f.key: f for f in FORMATS}


def available(fmt: Format) -> bool:
    """Ist das Format nutzbar, oder fehlt ein Paket?"""
    if not fmt.needs:
        return True
    try:
        __import__(fmt.needs)
        return True
    except ImportError:
        return False


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def suggest_filename(name: str, fmt_key: str) -> str:
    """Dateiname aus Auftragsname und Zeitstempel."""
    stem = "".join(c if c.isalnum() or c in " -_" else "_" for c in (name or "export"))
    stem = "_".join(stem.split())[:60] or "export"
    suffix = FORMAT_BY_KEY.get(fmt_key, FORMATS[0]).suffix
    return f"{stem}_{time.strftime('%Y-%m-%d_%H%M')}{suffix}"


# --------------------------------------------------------------------------
def to_csv(rows: Rows, path: Path, meta: dict | None = None) -> None:
    columns = all_columns(rows)
    with open(path, "w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow({c: _stringify(row.get(c)) for c in columns})


def to_json(rows: Rows, path: Path, meta: dict | None = None) -> None:
    payload = {
        "erzeugt": time.strftime("%Y-%m-%d %H:%M:%S"),
        "anzahl": len(rows),
        "meta": meta or {},
        "daten": rows,
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), "utf-8")


def to_jsonl(rows: Rows, path: Path, meta: dict | None = None) -> None:
    with open(path, "w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False) + "\n")


def to_markdown(rows: Rows, path: Path, meta: dict | None = None) -> None:
    columns = all_columns(rows)
    lines = [
        f"# Scrape-Ergebnis",
        "",
        f"Erzeugt: {time.strftime('%d.%m.%Y %H:%M')}  |  Datensätze: {len(rows)}",
        "",
    ]
    if meta and meta.get("zusammenfassung"):
        lines += ["## Zusammenfassung", "", str(meta["zusammenfassung"]), ""]

    if columns:
        lines.append("| " + " | ".join(columns) + " |")
        lines.append("|" + "|".join(["---"] * len(columns)) + "|")
        for row in rows:
            cells = [_stringify(row.get(c)).replace("|", "\\|").replace("\n", " ")
                     for c in columns]
            lines.append("| " + " | ".join(cells) + " |")
    path.write_text("\n".join(lines), "utf-8")


def to_html(rows: Rows, path: Path, meta: dict | None = None) -> None:
    columns = all_columns(rows)
    head = "".join(f"<th>{escape(c)}</th>" for c in columns)
    body = "".join(
        "<tr>" + "".join(
            f"<td>{escape(_stringify(row.get(c)))}</td>" for c in columns
        ) + "</tr>"
        for row in rows
    )
    summary = ""
    if meta and meta.get("zusammenfassung"):
        summary = f'<div class="summary">{escape(str(meta["zusammenfassung"]))}</div>'

    path.write_text(
        f"""<!doctype html>
<html lang="de"><head><meta charset="utf-8">
<title>Scrape-Ergebnis</title>
<style>
  :root {{ color-scheme: light dark; }}
  body {{ font: 15px/1.55 system-ui, -apple-system, Segoe UI, sans-serif;
         margin: 0; padding: 32px; background: #f6f7f9; color: #16181d; }}
  @media (prefers-color-scheme: dark) {{
    body {{ background: #14151a; color: #e8e9ed; }}
    table, .summary {{ background: #1c1e25 !important; }}
    th {{ background: #262933 !important; }}
    td, th {{ border-color: #2f3341 !important; }}
  }}
  h1 {{ font-size: 22px; margin: 0 0 4px; }}
  .meta {{ color: #6b7280; font-size: 13px; margin-bottom: 20px; }}
  .summary {{ background: #fff; border-left: 3px solid #7c5cff; padding: 14px 18px;
             border-radius: 8px; margin-bottom: 20px; white-space: pre-wrap; }}
  .wrap {{ overflow-x: auto; border-radius: 10px; }}
  table {{ border-collapse: collapse; width: 100%; background: #fff; font-size: 14px; }}
  th, td {{ padding: 9px 12px; text-align: left; border-bottom: 1px solid #e5e7eb;
           vertical-align: top; max-width: 460px; }}
  th {{ background: #eef0f4; position: sticky; top: 0; font-weight: 600; }}
  tr:hover td {{ background: rgba(124,92,255,.07); }}
</style></head>
<body>
<h1>Scrape-Ergebnis</h1>
<div class="meta">{escape(time.strftime('%d.%m.%Y %H:%M'))} &middot; {len(rows)} Datensätze</div>
{summary}
<div class="wrap"><table><thead><tr>{head}</tr></thead><tbody>{body}</tbody></table></div>
</body></html>""",
        "utf-8",
    )


def to_txt(rows: Rows, path: Path, meta: dict | None = None) -> None:
    lines = [f"Scrape-Ergebnis - {time.strftime('%d.%m.%Y %H:%M')}",
             f"{len(rows)} Datensätze", "=" * 60, ""]
    if meta and meta.get("zusammenfassung"):
        lines += [str(meta["zusammenfassung"]), "", "=" * 60, ""]
    for index, row in enumerate(rows, 1):
        lines.append(f"[{index}]")
        for key, value in row.items():
            lines.append(f"  {key}: {_stringify(value)}")
        lines.append("")
    path.write_text("\n".join(lines), "utf-8")


def to_xlsx(rows: Rows, path: Path, meta: dict | None = None) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    columns = all_columns(rows)
    book = Workbook()
    sheet = book.active
    sheet.title = "Ergebnisse"

    header_fill = PatternFill("solid", fgColor="4B3FBF")
    header_font = Font(color="FFFFFF", bold=True)
    for index, column in enumerate(columns, 1):
        cell = sheet.cell(row=1, column=index, value=column)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(vertical="center")

    for r_index, row in enumerate(rows, 2):
        for c_index, column in enumerate(columns, 1):
            sheet.cell(row=r_index, column=c_index, value=_stringify(row.get(column))[:32000])

    for index, column in enumerate(columns, 1):
        width = max(len(column), *(len(_stringify(r.get(column))[:60]) for r in rows)) \
            if rows else len(column)
        sheet.column_dimensions[get_column_letter(index)].width = min(max(width + 2, 10), 60)

    sheet.freeze_panes = "A2"
    if columns:
        sheet.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"

    if meta:
        info = book.create_sheet("Info")
        info["A1"] = "Feld"
        info["B1"] = "Wert"
        info["A1"].font = info["B1"].font = Font(bold=True)
        for index, (key, value) in enumerate(meta.items(), 2):
            info.cell(row=index, column=1, value=str(key))
            info.cell(row=index, column=2, value=_stringify(value)[:32000])
        info.column_dimensions["A"].width = 24
        info.column_dimensions["B"].width = 80

    book.save(path)


def to_pdf(rows: Rows, path: Path, meta: dict | None = None) -> None:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle,
    )

    columns = all_columns(rows)[:8]  # mehr passt nicht lesbar aufs Blatt
    styles = getSampleStyleSheet()
    cell_style = styles["BodyText"].clone("cell")
    cell_style.fontSize = 7.5
    cell_style.leading = 9.5

    document = SimpleDocTemplate(
        str(path), pagesize=landscape(A4),
        leftMargin=12 * mm, rightMargin=12 * mm,
        topMargin=12 * mm, bottomMargin=12 * mm,
        title="Scrape-Ergebnis",
    )

    story: list[Any] = [
        Paragraph("Scrape-Ergebnis", styles["Title"]),
        Paragraph(
            f"{time.strftime('%d.%m.%Y %H:%M')} &middot; {len(rows)} Datens&auml;tze",
            styles["Normal"],
        ),
        Spacer(1, 8),
    ]
    if meta and meta.get("zusammenfassung"):
        story.append(Paragraph("<b>Zusammenfassung</b>", styles["Heading3"]))
        story.append(Paragraph(escape(str(meta["zusammenfassung"])).replace("\n", "<br/>"),
                               styles["BodyText"]))
        story.append(Spacer(1, 10))

    if columns:
        table_data = [[Paragraph(f"<b>{escape(c)}</b>", cell_style) for c in columns]]
        for row in rows[:2000]:  # sonst wird das Dokument unbrauchbar gross
            table_data.append([
                Paragraph(escape(_stringify(row.get(c))[:300]), cell_style) for c in columns
            ])
        table = Table(table_data, repeatRows=1, hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4B3FBF")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#C9CCD6")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1),
             [colors.white, colors.HexColor("#F2F1FB")]),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(table)
        if len(rows) > 2000:
            story.append(Spacer(1, 6))
            story.append(Paragraph(
                f"... {len(rows) - 2000} weitere Datens&auml;tze nicht abgedruckt.",
                styles["Italic"]))

    document.build(story)


def to_sqlite(rows: Rows, path: Path, meta: dict | None = None) -> None:
    columns = all_columns(rows)
    connection = sqlite3.connect(str(path))
    try:
        safe = [f'"{c}"' for c in columns]
        connection.execute("DROP TABLE IF EXISTS ergebnisse")
        connection.execute(
            f"CREATE TABLE ergebnisse ({', '.join(f'{c} TEXT' for c in safe) or 'leer TEXT'})"
        )
        if columns:
            placeholders = ",".join("?" * len(columns))
            connection.executemany(
                f"INSERT INTO ergebnisse ({', '.join(safe)}) VALUES ({placeholders})",
                [tuple(_stringify(row.get(c)) for c in columns) for row in rows],
            )
        connection.execute("CREATE TABLE IF NOT EXISTS info (feld TEXT, wert TEXT)")
        connection.executemany(
            "INSERT INTO info VALUES (?,?)",
            [(str(k), _stringify(v)) for k, v in (meta or {}).items()],
        )
        connection.commit()
    finally:
        connection.close()


WRITERS: dict[str, Callable[[Rows, Path, dict | None], None]] = {
    "csv": to_csv,
    "xlsx": to_xlsx,
    "json": to_json,
    "jsonl": to_jsonl,
    "markdown": to_markdown,
    "html": to_html,
    "pdf": to_pdf,
    "txt": to_txt,
    "sqlite": to_sqlite,
}


def export(rows: Rows, path: str | Path, fmt_key: str, meta: dict | None = None) -> Path:
    """Zeilen in eine Datei schreiben. Gibt den Pfad zurück."""
    writer = WRITERS.get(fmt_key)
    if writer is None:
        raise ValueError(f"Unbekanntes Format: {fmt_key}")

    fmt = FORMAT_BY_KEY[fmt_key]
    if not available(fmt):
        raise RuntimeError(
            f"Für {fmt.label} fehlt das Paket '{fmt.needs}'. "
            f"Installieren mit: pip install {fmt.needs}"
        )

    target = Path(path)
    target.parent.mkdir(parents=True, exist_ok=True)
    writer(rows, target, meta)
    return target


def to_clipboard_text(rows: Rows, separator: str = "\t") -> str:
    """Tabelle als Text für die Zwischenablage (fügt sich in Excel ein)."""
    columns = all_columns(rows)
    lines = [separator.join(columns)]
    for row in rows:
        lines.append(separator.join(
            _stringify(row.get(c)).replace("\n", " ").replace(separator, " ")
            for c in columns
        ))
    return "\n".join(lines)
